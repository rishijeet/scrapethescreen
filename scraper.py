## Author Rishijeet Mishra
## Change the company URL for now, will enhance to use to be used as param

import pandas as pd
import requests
from bs4 import BeautifulSoup
import logging
import sys
import re
import numpy as np
from sklearn.linear_model import LinearRegression
from openpyxl.styles import PatternFill


# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def get_soup(url: str):
    """
    Fetches the webpage content and returns a BeautifulSoup object.
    
    Args:
        url: The URL of the webpage to scrape.
        
    Returns:
        A BeautifulSoup object or None if the request fails.
    """
    try:
        # Using a User-Agent to mimic a browser
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }
        response = requests.get(url, headers=headers, timeout=10)
        response.raise_for_status()  # Raises an HTTPError for bad responses (4xx or 5xx)
        return BeautifulSoup(response.text, 'html.parser')
    except requests.exceptions.RequestException as e:
        logging.error(f"Error fetching URL {url}: {e}")
        return None

def scrape_table_to_dataframe(soup: BeautifulSoup, section_id: str) -> pd.DataFrame | None:
    """
    Finds a table within a specific section by its ID and scrapes it into a pandas DataFrame.

    Args:
        soup: The BeautifulSoup object of the page.
        section_id: The 'id' attribute of the <section> tag containing the table.

    Returns:
        A pandas DataFrame containing the scraped data, or None if the table is not found.
    """
    section = soup.find('section', id=section_id)
    if not section:
        logging.warning(f"Could not find section with id: {section_id}")
        return None

    table = section.find('table', class_='data-table')
    if not table:
        logging.warning(f"Could not find a data table in section: {section_id}")
        return None

    # Extract headers
    headers = [th.get_text(strip=True) for th in table.find('thead').find_all('th')]

    # Extract rows
    data_rows = []
    for row in table.find('tbody').find_all('tr'):
        cols = [td.get_text(strip=True) for td in row.find_all('td')]
        # Clean the metric name (first column) by removing any trailing '+'
        # which is a UI element on the website for expandable rows.
        if cols:
            cols[0] = cols[0].removesuffix('+').strip()
        data_rows.append(cols)

    if not data_rows:
        logging.warning(f"No data rows found in table for section: {section_id}")
        return None

    # Create DataFrame
    df = pd.DataFrame(data_rows, columns=headers)
    return df

def clean_and_convert_to_numeric(series: pd.Series) -> pd.Series:
    """
    Cleans a pandas Series by removing commas and converting it to a numeric type.
    Non-numeric values are converted to NaN.
    """
    # Remove commas and strip whitespace
    cleaned_series = series.str.replace(',', '', regex=False).str.strip()
    # Convert to numeric, coercing errors (like '-') to NaN (Not a Number)
    return pd.to_numeric(cleaned_series, errors='coerce')

def forecast_metric(df: pd.DataFrame, metric_name: str, num_forecast_periods: int = 3) -> dict | None:
    """
    Forecasts a specific metric from a DataFrame using Linear Regression.

    Args:
        df: The DataFrame containing the financial data.
        metric_name: The name of the metric (row) to forecast.
        num_forecast_periods: The number of future periods to predict.

    Returns:
        A dictionary with last actual value and forecasted values, or None.
    """
    try:
        # Find the row for the specified metric
        metric_row = df[df.iloc[:, 0] == metric_name]
        if metric_row.empty:
            logging.warning(f"Metric '{metric_name}' not found.")
            return None

        # Get the data, drop the metric name column
        time_series_data = metric_row.iloc[0, 1:]
        numeric_data = clean_and_convert_to_numeric(time_series_data)

        # Prepare data for regression
        # X: Year periods, y: Metric values
        valid_data = numeric_data.dropna()
        if len(valid_data) < 2:
            logging.warning(f"Not enough data points for '{metric_name}' to forecast.")
            return None

        # Extract years from column headers like 'Mar 2023'
        year_headers = [col for col in valid_data.index if re.search(r'\d{4}', col)]
        if len(year_headers) != len(valid_data):
            logging.warning(f"Could not parse years for all data points in '{metric_name}'. Skipping forecast.")
            return None
        
        years = np.array([int(re.search(r'\d{4}', col).group()) for col in year_headers]).reshape(-1, 1)
        values = valid_data.values

        # Train the model
        model = LinearRegression()
        model.fit(years, values)

        # Predict future values
        last_year = years[-1][0]
        future_years = np.array([last_year + i for i in range(1, num_forecast_periods + 1)]).reshape(-1, 1)
        forecasted_values = model.predict(future_years)

        return {
            "last_actual_value": values[-1],
            "forecasts": list(forecasted_values),
            "forecast_years": [f"FY{yr[0]}" for yr in future_years]
        }
    except Exception as e:
        logging.error(f"Error during forecasting for '{metric_name}': {e}")
        return None

def get_rag_status(last_value: float, forecasts: list) -> str:
    """
    Determines the RAG status based on the trend of forecasted values.
    """
    if not forecasts or last_value is None or np.isnan(last_value) or last_value == 0:
        return "Amber" # Cannot determine trend

    avg_forecast = np.mean(forecasts)
    
    # Green: Forecast shows clear growth (>5% increase)
    if avg_forecast > last_value * 1.05:
        return "Green"
    # Red: Forecast shows clear decline (>5% decrease)
    elif avg_forecast < last_value * 0.95:
        return "Red"
    # Amber: Trend is stable or unclear
    else:
        return "Amber"

def main():
    """
    Main function to orchestrate the scraping and Excel file creation.
    """
    company_url = "https://www.screener.in/company/HDFCBANK/consolidated/"
    output_filename = "HDFCBANK_consolidated_financials.xlsx"

    logging.info(f"Starting scrape for {company_url}")
    soup = get_soup(company_url)
    if not soup:
        sys.exit(1) # Exit if page fetch failed

    # List of section IDs for the tables we want to scrape
    table_sections = [
        'quarters',
        'profit-loss',
        'balance-sheet',
        'cash-flow',
        'ratios'
    ]

    # Define which metrics to forecast from which tables
    metrics_to_forecast = {
        'profit-loss': ['Sales', 'Expenses', 'Net Profit'],
        'balance-sheet': ['Total Assets', 'Total Liabilities'],
        'cash-flow': [
            'Cash from Operating Activity',
            'Cash from Investing Activity',
            'Cash from Financing Activity',
            'Net Cash Flow'
        ]
    }
    summary_data = []

    try:
        with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
            logging.info(f"Writing data to {output_filename}")
            for section_id in table_sections:
                logging.info(f"Scraping section: {section_id}...")
                df_original = scrape_table_to_dataframe(soup, section_id)
                if df_original is not None:
                    # Use a clean sheet name from the section ID
                    sheet_name = section_id.replace('-', ' ').title()
                    df_original.to_excel(writer, sheet_name=sheet_name, index=False)
                    logging.info(f"Successfully wrote '{sheet_name}' sheet.")

                    # --- FORECASTING LOGIC ---
                    if section_id in metrics_to_forecast:
                        logging.info(f"--- Generating forecasts for {sheet_name} ---")
                        for metric in metrics_to_forecast[section_id]:
                            forecast_result = forecast_metric(df_original, metric)
                            if forecast_result:
                                rag_status = get_rag_status(forecast_result['last_actual_value'], forecast_result['forecasts'])
                                summary_row = {
                                    'Metric': metric,
                                    'Source Table': sheet_name,
                                    'Last Actual Value': f"{forecast_result['last_actual_value']:,.2f}",
                                    forecast_result['forecast_years'][0]: f"{forecast_result['forecasts'][0]:,.2f}",
                                    forecast_result['forecast_years'][1]: f"{forecast_result['forecasts'][1]:,.2f}",
                                    forecast_result['forecast_years'][2]: f"{forecast_result['forecasts'][2]:,.2f}",
                                    'Trend (RAG)': rag_status
                                }
                                summary_data.append(summary_row)
                                logging.info(f"Forecast for '{metric}' completed. Trend: {rag_status}")
                else:
                    logging.warning(f"Skipping section {section_id} as no data was found.")
            
            # --- WRITE SUMMARY SHEET ---
            if summary_data:
                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name="Forecast Summary", index=False)
                logging.info("Successfully wrote 'Forecast Summary' sheet.")

                # --- ADD CONDITIONAL FORMATTING ---
                logging.info("Applying conditional formatting to summary sheet...")
                try:
                    # Get the workbook and the summary worksheet objects
                    workbook = writer.book
                    worksheet = writer.sheets['Forecast Summary']

                    # Define color fills for RAG status
                    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    amber_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

                    # Find the column index for 'Trend (RAG)' (1-based for openpyxl)
                    rag_col_idx = summary_df.columns.get_loc('Trend (RAG)') + 1

                    # Iterate over the cells in the RAG column (skip the header) and apply formatting
                    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=rag_col_idx, max_col=rag_col_idx):
                        cell = row[0]
                        if cell.value == "Green":
                            cell.fill = green_fill
                        elif cell.value == "Amber":
                            cell.fill = amber_fill
                        elif cell.value == "Red":
                            cell.fill = red_fill
                    logging.info("Conditional formatting applied successfully.")
                except (KeyError, Exception) as e:
                    logging.warning(f"Could not apply conditional formatting. Reason: {e}")

        logging.info(f"Excel file '{output_filename}' created successfully.")
    except Exception as e:
        logging.error(f"An error occurred while writing the Excel file: {e}")

if __name__ == "__main__":
    main()